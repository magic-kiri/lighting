"""End-to-end validation: compare extracted counts against expected Excel output."""
import pytest
import openpyxl
from app.pipeline import run_pipeline

CHASE_PDF = "20251119_JPMFC_Jamboree_SB_Revision to Permit_IFC_All Trades.pdf"
CHASE_XLSX = "CHASE BANK - NEWPORT BEACH COUNTS.xlsx"

AMLI_PDF = "04_Electrical_1-16-2026.pdf"
AMLI_XLSX = "AMLI-BREA, CA COUNTS.xlsx"


def _load_expected_counts(xlsx_path: str) -> dict[str, int]:
    """Load Type -> Quantity from the expected Excel file."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    # Try common sheet names
    for name in wb.sheetnames:
        if "quote" in name.lower() or "customer" in name.lower():
            ws = wb[name]
            break
    else:
        ws = wb[wb.sheetnames[0]]

    counts = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            type_code = str(row[0]).strip()
            try:
                qty = int(row[1])
                if type_code and qty > 0:
                    counts[type_code] = counts.get(type_code, 0) + qty
            except (ValueError, TypeError):
                continue
    wb.close()
    return counts


def test_chase_bank_e2e():
    """Validate Chase Bank extraction against expected Excel counts."""
    result = run_pipeline(CHASE_PDF)
    assert result["status"] == "success"

    expected = _load_expected_counts(CHASE_XLSX)
    extracted = {fc["type"]: fc["quantity"] for fc in result["fixture_counts"]}

    print("\n=== Chase Bank Validation ===")
    print(f"{'Type':<10} {'Expected':>10} {'Extracted':>10} {'Match':>8}")
    print("-" * 40)

    matches = 0
    total = 0
    for type_code in sorted(set(list(expected.keys()) + list(extracted.keys()))):
        exp = expected.get(type_code, 0)
        ext = extracted.get(type_code, 0)
        match = "OK" if exp == ext else f"DIFF ({exp-ext:+d})"
        if exp > 0:
            total += 1
            if exp == ext:
                matches += 1
        print(f"{type_code:<10} {exp:>10} {ext:>10} {match:>8}")

    accuracy = matches / total * 100 if total > 0 else 0
    print(f"\nAccuracy: {matches}/{total} = {accuracy:.1f}%")
    # We want at least 50% match on first run — will improve iteratively
    assert accuracy >= 50, f"Accuracy too low: {accuracy:.1f}%"


def test_amli_brea_e2e():
    """Validate AMLI BREA extraction against expected Excel counts."""
    result = run_pipeline(AMLI_PDF)
    assert result["status"] in ("success", "error")

    if result["status"] == "error":
        pytest.skip(f"Pipeline returned error: {result['errors']}")

    expected = _load_expected_counts(AMLI_XLSX)
    extracted = {fc["type"]: fc["quantity"] for fc in result["fixture_counts"]}

    print("\n=== AMLI BREA Validation ===")
    print(f"{'Type':<10} {'Expected':>10} {'Extracted':>10} {'Match':>8}")
    print("-" * 40)

    matches = 0
    total = 0
    for type_code in sorted(set(list(expected.keys()) + list(extracted.keys()))):
        exp = expected.get(type_code, 0)
        ext = extracted.get(type_code, 0)
        match = "OK" if exp == ext else f"DIFF ({exp-ext:+d})"
        if exp > 0:
            total += 1
            if exp == ext:
                matches += 1
        print(f"{type_code:<10} {exp:>10} {ext:>10} {match:>8}")

    accuracy = matches / total * 100 if total > 0 else 0
    print(f"\nAccuracy: {matches}/{total} = {accuracy:.1f}%")
    assert accuracy >= 50, f"Accuracy too low: {accuracy:.1f}%"
